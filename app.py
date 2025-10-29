from flask import Flask, request, jsonify, send_file, session
import sqlite3, threading, os, time
from openpyxl import load_workbook
from watchdog.observers import Observer
from watchdog.events import FileSystemEventHandler

app = Flask(__name__)
app.secret_key = "halloween_secret_key_2025"

DB_NAME = 'students.db'
EXCEL_FILE = 'members.xlsx'
CODE_FILE = 'codes.xlsx'

print("현재 Flask 실행 경로:", os.getcwd())

# === 캐시 차단 ===
@app.after_request
def no_cache(response):
    response.headers["Cache-Control"] = "no-cache, no-store, must-revalidate"
    response.headers["Pragma"] = "no-cache"
    response.headers["Expires"] = "0"
    return response

# === DB 초기화 ===
def ensure_db():
    conn = sqlite3.connect(DB_NAME)
    cur = conn.cursor()

    cur.execute('''CREATE TABLE IF NOT EXISTS members (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        name TEXT NOT NULL,
        student_id TEXT NOT NULL UNIQUE
    )''')

    cur.execute('''CREATE TABLE IF NOT EXISTS attempts (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        student_id TEXT NOT NULL,
        code TEXT NOT NULL,
        timestamp DATETIME DEFAULT CURRENT_TIMESTAMP
    )''')

    cur.execute('''CREATE TABLE IF NOT EXISTS used_codes (
        code TEXT PRIMARY KEY,
        student_id TEXT NOT NULL,
        name TEXT NOT NULL,
        timestamp DATETIME DEFAULT CURRENT_TIMESTAMP
    )''')

    conn.commit()
    conn.close()

# === Excel → DB 업데이트 ===
def update_database_from_excel():
    if not os.path.exists(EXCEL_FILE):
        print("⚠️ Excel 파일이 없습니다.")
        return
    wb = load_workbook(EXCEL_FILE, data_only=True)
    ws = wb.active
    rows = []
    for r in range(2, ws.max_row + 1):
        name = ws.cell(r, 1).value
        sid = ws.cell(r, 2).value
        if name and sid:
            rows.append((str(name).strip(), str(sid).strip()))
    conn = sqlite3.connect(DB_NAME)
    cur = conn.cursor()
    cur.execute("DELETE FROM members")
    cur.executemany("INSERT OR REPLACE INTO members (name, student_id) VALUES (?, ?)", rows)
    conn.commit()
    conn.close()
    print(f"✅ Excel → DB 업데이트 완료! (총 {len(rows)}명)")

# === Excel 변경 감시 ===
class ExcelHandler(FileSystemEventHandler):
    def on_modified(self, event):
        if os.path.basename(event.src_path) == EXCEL_FILE:
            print("📂 members.xlsx 변경 감지 → DB 갱신")
            time.sleep(0.5)
            update_database_from_excel()

def start_watch():
    observer = Observer()
    observer.schedule(ExcelHandler(), ".", recursive=False)
    observer.start()
    print("👀 Excel 변경 감시 중...")
    while True:
        time.sleep(1)

# === 로그인 ===
@app.route('/login', methods=['POST'])
def login():
    data = request.get_json(force=True)
    name = data.get('name', '').strip()
    sid = data.get('student_id', '').strip()
    conn = sqlite3.connect(DB_NAME)
    cur = conn.cursor()
    cur.execute("SELECT 1 FROM members WHERE name=? AND student_id=?", (name, sid))
    result = cur.fetchone()
    conn.close()
    if result:
        session['name'] = name
        session['student_id'] = sid
        print(f"✅ 로그인 성공: {name} ({sid})")
        return jsonify({"success": True})
    else:
        print(f"❌ 로그인 실패: {name} ({sid})")
        return jsonify({"success": False, "message": "등록되지 않은 학생입니다."})

# === 코드 확인 ===
@app.route('/check_code', methods=['POST'])
def check_code():
    if 'student_id' not in session:
        return jsonify({"success": False, "message": "로그인 정보가 없습니다. 다시 로그인 해주세요."})

    sid = session['student_id']
    name = session.get('name', '')

    if not os.path.exists(CODE_FILE):
        return jsonify({"success": False, "message": "코드 데이터 파일이 없습니다."})

    data = request.get_json(force=True)
    code = data.get('code', '').strip()

    wb = load_workbook(CODE_FILE, data_only=True)
    ws = wb.active

    matched_page = None
    for r in range(2, ws.max_row + 1):
        c = str(ws.cell(r, 1).value).strip()
        page = str(ws.cell(r, 2).value).strip()
        if code == c:
            matched_page = page
            break

    if not matched_page:
        print(f"❌ 잘못된 코드: {code}")
        return jsonify({"success": False, "message": "존재하지 않는 코드입니다."})

    conn = sqlite3.connect(DB_NAME)
    cur = conn.cursor()

    if matched_page == "fake_ghost.html":
        print(f"😈 {name} ({sid}) fake_ghost 입력 - 제한 없음")
        conn.close()
        return jsonify({"success": True, "page": matched_page})

    cur.execute("SELECT student_id, name FROM used_codes WHERE code=?", (code,))
    already = cur.fetchone()
    if already and already[0] != sid:
        print(f"⚠️ 코드 {code}는 이미 {already[1]}({already[0]})가 사용함")
        conn.close()
        return jsonify({"success": False, "message": "이미 가져간 코드입니다."})

    cur.execute("SELECT COUNT(*) FROM attempts WHERE student_id=?", (sid,))
    count = cur.fetchone()[0]

    if count >= 3:
        print(f"🚫 {name} ({sid}) 코드 입력 초과 ({count}회)")
        conn.close()
        return jsonify({"success": False, "message": "최대 3회까지만 입력할 수 있습니다."})

    cur.execute("INSERT INTO attempts (student_id, code) VALUES (?, ?)", (sid, code))
    cur.execute("INSERT OR IGNORE INTO used_codes (code, student_id, name) VALUES (?, ?, ?)", (code, sid, name))

    conn.commit()
    conn.close()

    print(f"✅ {name} ({sid}) 코드 {code} 사용 → {matched_page}")
    return jsonify({"success": True, "page": matched_page})

# === 관리자 페이지 ===
@app.route('/admin')
def admin_page():
    return send_file("admin.html")

@app.route('/admin_data')
def admin_data():
    conn = sqlite3.connect(DB_NAME)
    cur = conn.cursor()
    cur.execute("SELECT student_id, code, timestamp FROM attempts ORDER BY timestamp DESC")
    attempts = [{"student_id": r[0], "code": r[1], "timestamp": r[2]} for r in cur.fetchall()]
    cur.execute("SELECT code, student_id, name, timestamp FROM used_codes ORDER BY timestamp DESC")
    used = [{"code": r[0], "student_id": r[1], "name": r[2], "timestamp": r[3]} for r in cur.fetchall()]
    conn.close()

    # ✅ 상품명 자동 매칭 추가
    product_map = {
        "boss_ghost.html": "상품권",
        "baby_ghost.html": "간식",
        "photo_ghost.html": "사진"
    }

    # codes.xlsx에서 페이지 불러와 매칭
    if os.path.exists(CODE_FILE):
        wb = load_workbook(CODE_FILE, data_only=True)
        ws = wb.active
        code_to_page = {str(ws.cell(r, 1).value).strip(): str(ws.cell(r, 2).value).strip() for r in range(2, ws.max_row + 1)}
        for u in used:
            page = code_to_page.get(u["code"])
            u["product"] = product_map.get(page, "-")

    return jsonify({"attempts": attempts, "used_codes": used})

@app.route('/reset_attempts', methods=['POST'])
def reset_attempts():
    conn = sqlite3.connect(DB_NAME)
    cur = conn.cursor()
    cur.execute("DELETE FROM attempts")
    conn.commit()
    conn.close()
    print("🧹 모든 시도 이력 삭제 완료")
    return jsonify({"success": True})

@app.route('/reset_used_codes', methods=['POST'])
def reset_used_codes():
    conn = sqlite3.connect(DB_NAME)
    cur = conn.cursor()
    cur.execute("DELETE FROM used_codes")
    conn.commit()
    conn.close()
    print("🧹 모든 사용된 코드 기록 삭제 완료")
    return jsonify({"success": True})

@app.route('/delete_code', methods=['POST'])
def delete_code():
    data = request.get_json(force=True)
    code = data.get('code', '').strip()

    if not code:
        return jsonify({"success": False, "message": "삭제할 코드가 지정되지 않았습니다."})

    conn = sqlite3.connect(DB_NAME)
    cur = conn.cursor()
    cur.execute("DELETE FROM used_codes WHERE code=?", (code,))
    deleted = cur.rowcount
    conn.commit()
    conn.close()

    if deleted > 0:
        print(f"🗑️ 관리자: 코드 {code} 삭제 완료")
        return jsonify({"success": True})
    else:
        print(f"⚠️ 관리자: 코드 {code} 삭제 실패 (존재하지 않음)")
        return jsonify({"success": False, "message": "존재하지 않는 코드입니다."})

# === 라우트 ===
@app.route('/')
def root():
    return send_file("login.html")

@app.route('/login_success.html')
def success_page():
    return send_file("login_success.html")

@app.route('/<path:filename>')
def serve_any(filename):
    if os.path.exists(filename):
        return send_file(filename)
    else:
        return f"⚠️ 파일을 찾을 수 없습니다: {filename}", 404

# === 실행 (Render 호환) ===
if __name__ == '__main__':
    ensure_db()
    update_database_from_excel()
    threading.Thread(target=start_watch, daemon=True).start()

    port = int(os.environ.get("PORT", 5000))
    app.run(host="0.0.0.0", port=port)
